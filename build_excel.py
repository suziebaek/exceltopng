"""
grammar_content.xlsx 생성 스크립트
- rule_reminder : 규칙 요약 카드
- error_spotlight : 오류 패턴 스포트라이트
- practice_ref : 교재 연습 문항 및 정답
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="B9A6DE", end_color="B9A6DE", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=11, bold=True)
BODY_FONT = Font(name="Arial", size=11)
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="center", horizontal="left")
WRAP_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")


def style_sheet(ws, headers, rows, col_widths, center_cols=None):
    center_cols = center_cols or []
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP_CENTER
        cell.border = BORDER

    for r, row in enumerate(rows, start=2):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = BODY_FONT
            cell.alignment = WRAP_CENTER if c in center_cols else WRAP
            cell.border = BORDER

    for c, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.row_dimensions[1].height = 22
    for r in range(2, len(rows) + 2):
        ws.row_dimensions[r].height = 60


wb = openpyxl.Workbook()

# ---------- 1. rule_reminder ----------
ws1 = wb.active
ws1.title = "rule_reminder"
headers1 = ["구성 요소", "내용"]
rows1 = [
    ["핵심 규칙①",
     "주어 + 자동사 (+ 수식어)\n"
     "My mom and I arrived at Beecher Prep.\n"
     "Mr. Tushman was waiting for us."],
    ["핵심 규칙②",
     "주어 + 자동사 + 보어(형용사/명사)\n"
     "He seemed nice.\n"
     "He looked different."],
    ["주의 패턴",
     "⚠ 감각동사 뒤 → 형용사 (부사 X)\n"
     "He seemed nicely. (X) → He seemed nice. (✓)"],
]
style_sheet(ws1, headers1, rows1, col_widths=[16, 70], center_cols=[1])

# ---------- 2. error_spotlight ----------
ws2 = wb.create_sheet("error_spotlight")
headers2 = ["오류 패턴", "틀린 문장", "올바른 문장", "이유"]
rows2 = [
    ["감각동사 + 부사", "The fresh air smells nicely.", "The fresh air smells nice.",
     "감각동사는 형용사 보어"],
    ["자동사 + 목적어 직결", "He arrived the school early.", "He arrived at the school early.",
     "arrive는 자동사, 전치사 필요"],
]
style_sheet(ws2, headers2, rows2, col_widths=[20, 28, 30, 26], center_cols=[1, 4])

# ---------- 3. practice_ref ----------
ws3 = wb.create_sheet("practice_ref")
headers3 = ["번호", "내용", "정답 및 해설"]
rows3 = [
    ["01", "He seems [ nervous / nervously ].",
     "정답: nervous\nseem은 연결동사이므로 주어의 상태를 설명 → 형용사 사용"],
    ["02", "The fresh air in the forest smells [ nice / nicely ].",
     "정답: nice\nsmell은 감각동사(연결동사처럼 사용) → 상태 설명 → 형용사 사용"],
]
style_sheet(ws3, headers3, rows3, col_widths=[8, 46, 55], center_cols=[1])

wb.save("/home/claude/grammar_app/grammar_content.xlsx")
print("saved")
